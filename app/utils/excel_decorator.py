from functools import wraps
from openpyxl import Workbook
from fastapi.responses import FileResponse


def export_excel(sheet_name, headers, file_name):
    def decorator(func):

        @wraps(func)
        def wrapper(*args, **kwargs):

            # Call the original service
            data = func(*args, **kwargs)

            wb = Workbook()
            sheet = wb.active
            sheet.title = sheet_name

            # Write headers
            for col, header in enumerate(headers, start=1):
                sheet.cell(row=1, column=col).value = header

            # Write data
            row = 2
            for record in data:
                for col, value in enumerate(record, start=1):
                    sheet.cell(row=row, column=col).value = value
                row += 1

            wb.save(file_name)

            return FileResponse(
                path=file_name,
                filename=file_name,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        return wrapper

    return decorator